#!/usr/bin/env python3
"""
Node Cross-Reference Tool
=========================

This tool cross-references tickets from a CSV file with an official offline nodes report
to determine which tickets can be closed.

Usage:
    python node_cross_reference.py

Input Files:
    - nodes_tickets.csv: CSV with ticket information
    - nodes_report.txt: Official offline nodes report

Output:
    - results_can_close.csv: Tickets that can be definitely closed
    - results_need_review.csv: Tickets that need additional review
    - summary_report.txt: Summary of the analysis
"""

import csv
import re
from typing import Dict, List, Set, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
import os
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Import CSV auto-repair functionality
try:
    from csv_auto_repair import CSVRepairer, cleanup_temp_file
    CSV_REPAIR_AVAILABLE = True
except ImportError:
    CSV_REPAIR_AVAILABLE = False


@dataclass
class Ticket:
    """Represents a ticket from the CSV file"""
    site: str
    number: str
    description: str
    priority: str
    created: str
    updated: str
    resolved: Optional[str] = None
    assignment_group: Optional[str] = None
    store_number: Optional[int] = None
    node_number: Optional[int] = None
    
    @property
    def is_closed(self) -> bool:
        """Check if ticket is closed based on resolved field"""
        return self.resolved is not None and self.resolved.strip() != ""
    
    @property
    def created_datetime(self) -> Optional[datetime]:
        """Parse created date to datetime object"""
        return self._parse_date(self.created)
    
    @property
    def updated_datetime(self) -> Optional[datetime]:
        """Parse updated date to datetime object"""
        return self._parse_date(self.updated)
    
    @property
    def resolved_datetime(self) -> Optional[datetime]:
        """Parse resolved date to datetime object"""
        if not self.resolved or not self.resolved.strip():
            return None
        return self._parse_date(self.resolved)
    
    def _parse_date(self, date_str: str) -> Optional[datetime]:
        """Parse date string with multiple format support"""
        if not date_str or not date_str.strip():
            return None
            
        date_str = date_str.strip()
        
        # Common date formats found in the CSV
        formats = [
            '%d-%b-%Y %H:%M:%S',  # 28-May-2025 13:25:39
            '%Y-%m-%d %H:%M:%S',  # 2025-05-28 13:25:39
            '%m/%d/%Y %H:%M:%S',  # 05/28/2025 13:25:39
            '%d/%m/%Y %H:%M:%S',  # 28/05/2025 13:25:39
            '%Y-%m-%d',           # 2025-05-28
            '%d-%b-%Y',           # 28-May-2025
            '%m/%d/%Y',           # 05/28/2025
            '%d/%m/%Y',           # 28/05/2025
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        # If all formats fail, return None
        print(f"Warning: Could not parse date '{date_str}'")
        return None
    
    def is_reopenable(self, max_days: int = 7) -> bool:
        """Check if ticket can be reopened (closed within max_days)"""
        if not self.is_closed:
            return False  # Already open
            
        resolved_dt = self.resolved_datetime
        if resolved_dt is None:
            return False  # Can't determine resolved date
            
        days_since_closed = (datetime.now() - resolved_dt).days
        return days_since_closed <= max_days


@dataclass
class OfflineNode:
    """Represents an offline node from the report"""
    store_number: int
    node_number: int
    esp_id: str
    last_seen: str
    
    @property
    def last_seen_datetime(self) -> Optional[datetime]:
        """Parse last seen date to datetime object"""
        return self._parse_date(self.last_seen)
    
    def _parse_date(self, date_str: str) -> Optional[datetime]:
        """Parse date string with multiple format support"""
        if not date_str or not date_str.strip():
            return None
            
        date_str = date_str.strip()
        
        # Common date formats in node report
        formats = [
            '%Y-%m-%d %H:%M:%S',  # 2025-07-12 14:54:14
            '%d-%b-%Y %H:%M:%S',  # 12-Jul-2025 14:54:14
            '%m/%d/%Y %H:%M:%S',  # 07/12/2025 14:54:14
            '%d/%m/%Y %H:%M:%S',  # 12/07/2025 14:54:14
            '%Y-%m-%d',           # 2025-07-12
            '%d-%b-%Y',           # 12-Jul-2025
            '%m/%d/%Y',           # 07/12/2025
            '%d/%m/%Y',           # 12/07/2025
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        # If all formats fail, return None
        print(f"Warning: Could not parse node last seen date '{date_str}'")
        return None
    
    def days_offline(self) -> Optional[int]:
        """Calculate how many days the node has been offline"""
        last_seen_dt = self.last_seen_datetime
        if last_seen_dt is None:
            return None
        return (datetime.now() - last_seen_dt).days
    
    def is_long_term_offline(self, threshold_days: int = 30) -> bool:
        """Check if node has been offline for a long time"""
        days = self.days_offline()
        return days is not None and days >= threshold_days


@dataclass
class AnalysisResult:
    """Result of analyzing a ticket"""
    ticket: Ticket
    status: str  # "can_close", "needs_review", "suggest_reopen", "closed_ok", "closed_too_old", "error"
    reason: str
    store_in_report: bool = False
    node_in_report: bool = False
    confidence: str = "medium"  # "high", "medium", "low"
    business_logic_flag: str = ""  # Additional flags like "do_not_close", "workflow_status", etc.
    temporal_analysis: str = ""  # Temporal correlation insights
    days_offline: Optional[int] = None  # Days the node has been offline
    reopenable: bool = False  # Whether ticket can be reopened based on age


class NodeCrossReference:
    def __init__(self):
        self.tickets: List[Ticket] = []
        self.offline_nodes: Dict[int, Set[int]] = {}  # store_number -> set of offline node numbers
        self.offline_nodes_detailed: Dict[Tuple[int, int], OfflineNode] = {}  # (store, node) -> OfflineNode
        self.results: List[AnalysisResult] = []
        self.saf_stores: Set[int] = set()  # Stores with SAF markers
        self.both_nodes_offline_stores: Set[int] = set()  # Stores with both nodes offline
        self.stores_with_tickets: Set[int] = set()  # Stores that have tickets
        self.temp_csv_file: Optional[str] = None  # Track temporary repaired CSV file for cleanup

    def cleanup(self):
        """Clean up temporary files"""
        if self.temp_csv_file:
            cleanup_temp_file(self.temp_csv_file)
            self.temp_csv_file = None

    def __del__(self):
        """Destructor to ensure cleanup"""
        self.cleanup()

    def extract_store_number(self, site: str) -> Optional[int]:
        """Extract store number from site field - handles multiple Wendy's formats"""
        # Pattern 1: Wendy's #1234 or Wendys #1234 (with or without apostrophe)
        match = re.search(r"Wendy'?s\s*#(\d+)", site, re.IGNORECASE)
        if match:
            return int(match.group(1))

        # Pattern 2: WENDYS 04999-FZ-SW format (remove leading zeros)
        match = re.search(r"WENDYS\s+(\d+)(?:-[A-Z0-9-]+)?", site, re.IGNORECASE)
        if match:
            store_num_str = match.group(1).lstrip('0') or '0'  # Remove leading zeros, keep at least one digit
            return int(store_num_str)

        return None
    
    def extract_node_number(self, description: str) -> Optional[int]:
        """Extract node number from ticket description"""
        # Clean up description for easier parsing
        desc = description.upper()
        
        # Look for various node patterns
        patterns = [
            r'NODE\s*(\d+)',           # NODE1, NODE 1, etc.
            r'NODE\s*\((\d+)\)',       # NODE (1), NODE(2)
            r'NODE\s*#(\d+)',          # NODE#1, NODE #1
            r'\*\*NODE\s*(\d+)\*\*',   # **NODE1**
            r'ESP\s+NODE\s*(\d+)',     # ESP NODE 1
            r'NODE\(\s*(\d+)\s*\)',    # NODE(1), NODE( 2 )
            r'NODES\s*(\d+)',          # NODES1, NODES 2 (sometimes used)
            r'NODE-(\d+)',             # NODE-1
            r'NODE_(\d+)',             # NODE_1
        ]
        
        for pattern in patterns:
            match = re.search(pattern, desc)
            if match:
                node_num = int(match.group(1))
                # Validate node number (should be 1 or 2)
                if node_num in [1, 2]:
                    return node_num
                else:
                    # Invalid node number, treat as ambiguous
                    return None
        
        # If no specific node number found, check for generic "NODES" (ambiguous)
        if "NODES" in desc and "NODE" in desc:
            # Could be multiple nodes, return None to indicate review needed
            return None
        
        return None
    
    def detect_business_logic_flags(self, site: str, description: str) -> Tuple[bool, str]:
        """Detect business logic flags that should prevent auto-closing"""
        combined_text = f"{site} {description}".upper()
        
        # Check for explicit "do not close" instructions
        do_not_close_patterns = [
            r'DO\s*NOT\s*CLOSE',
            r'DON\'T\s*CLOSE',
            r'DONT\s*CLOSE',
            r'NOT\s*TO\s*CLOSE',
            r'KEEP\s*OPEN',
        ]
        
        for pattern in do_not_close_patterns:
            if re.search(pattern, combined_text):
                return True, "do_not_close"
        
        # Check for workflow status indicators
        workflow_patterns = [
            r'\*AEX\s+SUBMIT[TED]*\*',
            r'\*AWAITING\s+APPROVAL\*',
            r'\*TECH\s+SUBMIT[TED]*\*',
            r'\*APPROVED\*',
            r'\*AWAITING\s+INFO\*',
            r'\*AWAITING\s+ASSET\*',
            r'\*AWAITING\s+UPGRADE\*',
            r'\*EOL\*',
            r'WO\d+',  # Work Order references
            r'CS\d+',  # Case references in description
        ]
        
        for pattern in workflow_patterns:
            if re.search(pattern, combined_text):
                return True, "workflow_status"
        
        # Check for special instructions or notes
        special_instruction_patterns = [
            r'ONCE\s+\w+\s+NODE\s+IS\s+INSTALLED',
            r'AFTER\s+\w+',
            r'PENDING\s+\w+',
            r'WAITING\s+FOR\s+\w+',
        ]
        
        for pattern in special_instruction_patterns:
            if re.search(pattern, combined_text):
                return True, "special_instructions"
        
        return False, ""
    
    def analyze_temporal_correlation(self, ticket: Ticket) -> Tuple[str, Optional[int], bool]:
        """Analyze temporal correlation between ticket and node offline status"""
        temporal_analysis = ""
        days_offline = None
        timeline_issue = False
        
        if not ticket.store_number:
            return "No store number identified", days_offline, timeline_issue
            
        # Try to get specific offline node if we have node number
        if ticket.node_number is not None:
            node_key = (ticket.store_number, ticket.node_number)
            offline_node = self.offline_nodes_detailed.get(node_key)
            
            if offline_node is not None:
                return self._analyze_specific_node_temporal(ticket, offline_node)
        
        # If no specific node or node not found, provide general temporal analysis
        store_in_report = ticket.store_number in self.offline_nodes
        if not store_in_report:
            return "Store has no offline nodes currently", days_offline, timeline_issue
        
        # Get general store offline information
        offline_nodes_for_store = self.offline_nodes[ticket.store_number]
        
        # Try to find any detailed offline node info for this store
        store_offline_nodes = [
            self.offline_nodes_detailed.get((ticket.store_number, node_num))
            for node_num in offline_nodes_for_store
        ]
        store_offline_nodes = [node for node in store_offline_nodes if node is not None]
        
        if store_offline_nodes:
            # Use the node that's been offline the longest for temporal analysis
            longest_offline_node = max(store_offline_nodes, key=lambda x: x.days_offline())
            days_offline = longest_offline_node.days_offline()
            
            if len(offline_nodes_for_store) == 1:
                temporal_analysis = f"Single node offline for {days_offline} days"
            else:
                temporal_analysis = f"Multiple nodes offline (longest: {days_offline} days)"
        else:
            temporal_analysis = f"Store has {len(offline_nodes_for_store)} offline nodes (no detailed timing)"
            
        return temporal_analysis, days_offline, timeline_issue
    
    def _analyze_specific_node_temporal(self, ticket: Ticket, offline_node) -> Tuple[str, Optional[int], bool]:
        """Analyze temporal correlation for a specific identified offline node"""
        temporal_analysis = ""
        timeline_issue = False
            
        days_offline = offline_node.days_offline()
        
        ticket_created = ticket.created_datetime
        ticket_resolved = ticket.resolved_datetime
        node_last_seen = offline_node.last_seen_datetime
        
        if not all([ticket_created, node_last_seen]):
            return temporal_analysis, days_offline, timeline_issue
            
        # Calculate timeline relationships
        node_offline_before_ticket = node_last_seen < ticket_created
        
        if ticket_resolved and node_last_seen:
            node_offline_after_resolved = node_last_seen > ticket_resolved
            
            if node_offline_after_resolved and ticket.is_closed:
                temporal_analysis = f"Timeline anomaly: Node went offline AFTER ticket was resolved. Node offline since {offline_node.last_seen}"
                timeline_issue = True
            elif node_offline_before_ticket:
                days_before = (ticket_created - node_last_seen).days
                temporal_analysis = f"Node was offline {days_before} days before ticket was created"
            else:
                temporal_analysis = f"Normal timeline: Node offline before resolution"
        elif node_offline_before_ticket:
            days_before = (ticket_created - node_last_seen).days
            temporal_analysis = f"Node was offline {days_before} days before ticket was created"
        else:
            temporal_analysis = f"Node went offline around ticket creation time"
            
        return temporal_analysis, days_offline, timeline_issue
    
    def get_node_offline_duration(self, store_number: int, node_number: int) -> Optional[int]:
        """Get days a specific node has been offline"""
        node_key = (store_number, node_number)
        offline_node = self.offline_nodes_detailed.get(node_key)
        if offline_node:
            return offline_node.days_offline()
        return None
    
    def create_analysis_result(self, ticket: Ticket, status: str, reason: str, 
                             store_in_report: bool, node_in_report: bool, 
                             confidence: str, business_flag: str = "",
                             temporal_analysis: str = "", days_offline: Optional[int] = None) -> AnalysisResult:
        """Helper method to create AnalysisResult with all temporal fields"""
        return AnalysisResult(
            ticket=ticket,
            status=status,
            reason=reason,
            store_in_report=store_in_report,
            node_in_report=node_in_report,
            confidence=confidence,
            business_logic_flag=business_flag,
            temporal_analysis=temporal_analysis,
            days_offline=days_offline,
            reopenable=ticket.is_reopenable() if ticket.is_closed else True
        )
    
    def _format_offline_nodes_message(self, offline_nodes: set, suffix: str = "") -> str:
        """Format offline nodes message clearly to avoid confusion"""
        if len(offline_nodes) == 1:
            node_num = list(offline_nodes)[0]
            return f"Node {node_num} is offline {suffix}".strip()
        else:
            nodes_list = sorted(offline_nodes)
            return f"Nodes {' and '.join(map(str, nodes_list))} are offline {suffix}".strip()
    
    def determine_confidence(self, ticket: Ticket, store_in_report: bool, 
                           node_in_report: bool, business_flag: str) -> str:
        """Determine confidence level for the decision"""
        
        # Low confidence conditions
        if business_flag:
            return "low"
        
        if ticket.store_number is None:
            return "low"
        
        if ticket.node_number is None and store_in_report:
            return "low"
        
        # High confidence conditions
        if not store_in_report:
            # No nodes from this store are offline - very confident
            return "high"
        
        if store_in_report and ticket.node_number is not None:
            # We have specific node info and can check exact node status - high confidence
            return "high"
        
        # Medium confidence for everything else
        return "medium"
    
    def load_tickets(self, csv_file: str):
        """Load tickets from CSV file with dynamic column detection and automatic repair"""
        # Clean up any previous temporary file
        if self.temp_csv_file:
            cleanup_temp_file(self.temp_csv_file)
            self.temp_csv_file = None

        # Attempt automatic CSV repair if available
        actual_csv_file = csv_file
        if CSV_REPAIR_AVAILABLE:
            try:
                import logging
                # Create logger for repair operations
                logger = logging.getLogger('node_cross_reference.csv_repair')
                repairer = CSVRepairer(logger)
                actual_csv_file = repairer.auto_repair_csv(csv_file)

                # Track if we're using a temporary repaired file
                if actual_csv_file != csv_file:
                    self.temp_csv_file = actual_csv_file
                    print(f"CSV automatically repaired: {os.path.basename(csv_file)} → using repaired version")

            except Exception as e:
                print(f"Warning: CSV auto-repair failed, using original file: {e}")
                actual_csv_file = csv_file

        try:
            with open(actual_csv_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                # Validate required columns
                required_columns = ['Site', 'Number', 'Short description', 'Priority', 'Created', 'Updated']
                missing_columns = [col for col in required_columns if col not in reader.fieldnames]
                if missing_columns:
                    raise ValueError(f"Missing required columns in CSV: {', '.join(missing_columns)}")
                
                # Check for optional columns
                has_resolved = 'Resolved' in reader.fieldnames
                has_assignment_group = 'Assignment Group' in reader.fieldnames or 'Assignment group' in reader.fieldnames
                
                assignment_group_col = None
                if 'Assignment Group' in reader.fieldnames:
                    assignment_group_col = 'Assignment Group'
                elif 'Assignment group' in reader.fieldnames:
                    assignment_group_col = 'Assignment group'
                
                if has_resolved:
                    print(f"Found 'Resolved' column - will analyze both open and closed tickets")
                else:
                    print(f"No 'Resolved' column found - treating all tickets as open")
                    
                if has_assignment_group:
                    print(f"Found '{assignment_group_col}' column - will include assignment group analysis")
                else:
                    print(f"No 'Assignment Group' column found - skipping assignment group analysis")
                
                for row in reader:
                    try:
                        ticket = Ticket(
                            site=row['Site'],
                            number=row['Number'],
                            description=row['Short description'],
                            priority=row['Priority'],
                            created=row['Created'],
                            updated=row['Updated'],
                            resolved=row.get('Resolved', '') if has_resolved else None,
                            assignment_group=row.get(assignment_group_col, '') if has_assignment_group else None
                        )
                        
                        # Extract store and node numbers
                        ticket.store_number = self.extract_store_number(ticket.site)
                        ticket.node_number = self.extract_node_number(ticket.description)
                        
                        # Track stores that have tickets
                        if ticket.store_number:
                            self.stores_with_tickets.add(ticket.store_number)
                        
                        self.tickets.append(ticket)
                        
                    except Exception as e:
                        print(f"Warning: Failed to parse ticket row {row.get('Number', 'unknown')}: {e}")
                        continue
        
        except FileNotFoundError:
            raise FileNotFoundError(f"CSV file '{csv_file}' not found")
        except Exception as e:
            raise Exception(f"Error loading CSV file '{csv_file}': {e}")

        open_tickets = len([t for t in self.tickets if not t.is_closed])
        closed_tickets = len([t for t in self.tickets if t.is_closed])
        print(f"Loaded {len(self.tickets)} tickets from {os.path.basename(csv_file)} ({open_tickets} open, {closed_tickets} closed)")
    
    def load_offline_nodes(self, report_file: str):
        """Load offline nodes from the report file"""
        try:
            with open(report_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            if not content.strip():
                raise ValueError("Report file is empty")
            
            # Track critical stores
            self.saf_stores = set()  # Stores with SAF markers
            self.both_nodes_offline_stores = set()  # Stores with both nodes offline
            
            # Parse store sections
            store_sections = re.split(r'^Store #(\d+)', content, flags=re.MULTILINE)
            
            if len(store_sections) < 2:
                raise ValueError("No store sections found in report file. Expected format: 'Store #<number>'")
            
            for i in range(1, len(store_sections), 2):
                try:
                    store_number = int(store_sections[i])
                    section_content = store_sections[i + 1]
                    
                    # Check for SAF marker
                    if '!!! SAF !!!' in section_content:
                        self.saf_stores.add(store_number)
                    
                    # Find all nodes in this section with detailed information
                    node_pattern = r'NODE\s+(esp\d+-l0([12]))\s+OFFLINE\.\s+Last seen:\s+(.+)'
                    node_matches = re.findall(node_pattern, section_content)
                    
                    if store_number not in self.offline_nodes:
                        self.offline_nodes[store_number] = set()
                    
                    for esp_id, node_num_str, last_seen in node_matches:
                        node_number = int(node_num_str)
                        self.offline_nodes[store_number].add(node_number)
                        
                        # Create detailed OfflineNode object
                        offline_node = OfflineNode(
                            store_number=store_number,
                            node_number=node_number,
                            esp_id=esp_id,
                            last_seen=last_seen.strip()
                        )
                        
                        self.offline_nodes_detailed[(store_number, node_number)] = offline_node
                    
                    # Fallback: if the above pattern didn't match, try simpler pattern
                    if not node_matches:
                        simple_matches = re.findall(r'esp\d+-l0([12])', section_content)
                        for node_match in simple_matches:
                            node_number = int(node_match)
                            self.offline_nodes[store_number].add(node_number)
                    
                    # Check if both nodes are offline
                    if len(self.offline_nodes[store_number]) >= 2:
                        self.both_nodes_offline_stores.add(store_number)
                        
                except ValueError as e:
                    print(f"Warning: Failed to parse store section: {e}")
                    continue
        
        except FileNotFoundError:
            raise FileNotFoundError(f"Report file '{report_file}' not found")
        except Exception as e:
            raise Exception(f"Error loading report file '{report_file}': {e}")
        
        total_stores = len(self.offline_nodes)
        total_nodes = sum(len(nodes) for nodes in self.offline_nodes.values())
        saf_count = len(self.saf_stores)
        both_nodes_count = len(self.both_nodes_offline_stores)
        
        print(f"Loaded {total_nodes} offline nodes from {total_stores} stores")
        print(f"CRITICAL: {saf_count} stores with SAF markers")
        print(f"CRITICAL: {both_nodes_count} stores with both nodes offline")
    
    def get_missing_tickets(self) -> List[dict]:
        """Identify stores in offline report that don't have tickets"""
        missing_tickets = []
        
        for store_number, offline_nodes in self.offline_nodes.items():
            if store_number not in self.stores_with_tickets:
                # This store has offline nodes but no tickets
                is_saf = store_number in self.saf_stores
                is_both_offline = store_number in self.both_nodes_offline_stores
                
                # Determine priority
                if is_saf:
                    priority = "CRITICAL - SAF"
                    urgency = "Immediate"
                elif is_both_offline:
                    priority = "CRITICAL - Both Nodes"
                    urgency = "Immediate" 
                elif len(offline_nodes) > 1:
                    priority = "High"
                    urgency = "High"
                else:
                    priority = "Medium"
                    urgency = "Medium"
                
                for node_number in sorted(offline_nodes):
                    missing_tickets.append({
                        'store_number': store_number,
                        'site': f"Wendy's #{store_number}",
                        'node_number': node_number,
                        'priority': priority,
                        'urgency': urgency,
                        'offline_nodes': sorted(offline_nodes),
                        'is_saf': is_saf,
                        'is_both_offline': is_both_offline,
                        'suggested_description': f"HW-BOH-P2P-ESP Node {node_number}-Offline",
                        'reason': f"Node {node_number} offline - no existing ticket found"
                    })
        
        return missing_tickets
    
    def analyze_ticket(self, ticket: Ticket) -> AnalysisResult:
        """Analyze a single ticket to determine if it can be closed"""
        
        # Check for business logic flags first
        has_business_flag, business_flag = self.detect_business_logic_flags(
            ticket.site, ticket.description
        )
        
        # Perform temporal correlation analysis FIRST - needed for all paths
        temporal_analysis, days_offline, has_timeline_issue = self.analyze_temporal_correlation(ticket)
        
        # TEMPORAL INTELLIGENCE: Check if closed ticket is too old to reopen
        if ticket.is_closed and not ticket.is_reopenable(max_days=7):
            return AnalysisResult(
                ticket=ticket,
                status="closed_too_old",
                reason=f"Ticket closed more than 7 days ago ({ticket.resolved}) - cannot be reopened",
                store_in_report=False,
                node_in_report=False,
                confidence="high",
                business_logic_flag=business_flag,
                temporal_analysis=temporal_analysis,
                days_offline=days_offline,
                reopenable=False
            )
        
        # Check if we could extract store number
        if ticket.store_number is None:
            confidence = self.determine_confidence(ticket, False, False, business_flag)
            return AnalysisResult(
                ticket=ticket,
                status="error",
                reason="Could not extract store number from site field",
                store_in_report=False,
                node_in_report=False,
                confidence=confidence,
                business_logic_flag=business_flag,
                temporal_analysis=temporal_analysis,
                days_offline=days_offline,
                reopenable=ticket.is_reopenable() if ticket.is_closed else True
            )
        
        # Check if store is in the offline report
        store_in_report = ticket.store_number in self.offline_nodes
        
        # Check for CRITICAL conditions - SAF or both nodes offline
        is_saf_store = ticket.store_number in self.saf_stores
        is_both_nodes_offline = ticket.store_number in self.both_nodes_offline_stores
        
        # CRITICAL CONDITION: SAF stores - NEVER auto-close
        if is_saf_store:
            offline_nodes_for_store = self.offline_nodes.get(ticket.store_number, set())
            return self.create_analysis_result(
                ticket=ticket,
                status="needs_review",
                reason=f"CRITICAL: Store has SAF (Store and Forward) failure - both nodes offline ({sorted(offline_nodes_for_store)}). REQUIRES IMMEDIATE ATTENTION.",
                store_in_report=True,
                node_in_report=True,  # SAF means both nodes are problematic
                confidence="high",  # High confidence this needs review
                business_flag="critical_saf",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        # CRITICAL CONDITION: Both nodes offline - NEVER auto-close  
        if is_both_nodes_offline:
            offline_nodes_for_store = self.offline_nodes[ticket.store_number]
            return self.create_analysis_result(
                ticket=ticket,
                status="needs_review", 
                reason=f"CRITICAL: Store has BOTH nodes offline ({sorted(offline_nodes_for_store)}). Complete store connectivity loss. REQUIRES IMMEDIATE ATTENTION.",
                store_in_report=True,
                node_in_report=True,  # Both nodes are offline
                confidence="high",  # High confidence this needs review
                business_flag="critical_both_nodes_offline",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        # If business logic flag is present, handle based on type and node status
        if has_business_flag:
            flag_descriptions = {
                "do_not_close": "Ticket contains 'do not close' instructions",
                "workflow_status": "Ticket has workflow status indicators", 
                "special_instructions": "Ticket contains special handling instructions",
                "critical_saf": "CRITICAL: Store has SAF (Store and Forward) failure",
                "critical_both_nodes_offline": "CRITICAL: Store has both nodes offline"
            }
            
            # SPECIAL CASE: Workflow status tickets can be closed if nodes are back online
            if business_flag == "workflow_status":
                # GREEN ZONE: If store is not in report at all, definitely close the ticket
                if not store_in_report:
                    confidence = self.determine_confidence(ticket, False, False, "")
                    return AnalysisResult(
                        ticket=ticket,
                        status="can_close",
                        reason="Workflow status ticket - store not in offline report, all nodes confirmed online",
                        store_in_report=False,
                        node_in_report=False,
                        confidence=confidence,
                        business_logic_flag=business_flag
                    )
                
                # Store is in report, check if specific node is actually offline
                offline_nodes_for_store = self.offline_nodes[ticket.store_number]
                if ticket.node_number is None:
                    # Can't identify specific node, but store has some offline nodes - needs review
                    node_actually_offline = True
                else:
                    # Check if the specific node mentioned in the ticket is offline
                    node_actually_offline = ticket.node_number in offline_nodes_for_store
                
                if not node_actually_offline:
                    # Specific node is back online, workflow status ticket can be closed
                    confidence = self.determine_confidence(ticket, store_in_report, False, "")
                    status_detail = f"Node {ticket.node_number} is back online (other " + self._format_offline_nodes_message(offline_nodes_for_store).lower() + ")"
                    
                    return AnalysisResult(
                        ticket=ticket,
                        status="can_close",
                        reason=f"Workflow status ticket - node is back online. Status: {status_detail}",
                        store_in_report=store_in_report,
                        node_in_report=False,
                        confidence=confidence,
                        business_logic_flag=business_flag
                    )
            
            # All other business logic flags or workflow status with nodes still offline - needs review
            confidence = self.determine_confidence(ticket, store_in_report, False, business_flag)
            base_reason = flag_descriptions.get(business_flag, 'Business logic flag detected')
            
            if not store_in_report:
                status_detail = "No nodes from this store are currently offline"
            else:
                offline_nodes_for_store = self.offline_nodes[ticket.store_number]
                if ticket.node_number is None:
                    if len(offline_nodes_for_store) == 1:
                        node_num = list(offline_nodes_for_store)[0]
                        status_detail = f"Node {node_num} is offline, but couldn't identify specific node from ticket"
                    else:
                        nodes_list = sorted(offline_nodes_for_store)
                        status_detail = f"Nodes {' and '.join(map(str, nodes_list))} are offline, but couldn't identify specific node from ticket"
                else:
                    node_in_report = ticket.node_number in offline_nodes_for_store
                    if node_in_report:
                        status_detail = f"Node {ticket.node_number} IS confirmed offline"
                    else:
                        if len(offline_nodes_for_store) == 1:
                            other_node = list(offline_nodes_for_store)[0]
                            status_detail = f"Node {ticket.node_number} is NOT offline (Node {other_node} is offline)"
                        else:
                            nodes_list = sorted(offline_nodes_for_store)
                            status_detail = f"Node {ticket.node_number} is NOT offline (Nodes {' and '.join(map(str, nodes_list))} are offline)"
            
            reason = f"{base_reason} - requires manual review. Status: {status_detail}"
            
            return self.create_analysis_result(
                ticket=ticket,
                status="needs_review",
                reason=reason,
                store_in_report=store_in_report,
                node_in_report=ticket.node_number in self.offline_nodes.get(ticket.store_number, set()) if ticket.node_number else False,
                confidence=confidence,
                business_flag=business_flag,
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        if not store_in_report:
            # Store not in report means no nodes from this store are currently offline
            confidence = self.determine_confidence(ticket, False, False, business_flag)
            return AnalysisResult(
                ticket=ticket,
                status="can_close",
                reason="Store not in offline report - no nodes from this store are currently offline",
                store_in_report=False,
                node_in_report=False,
                confidence=confidence,
                business_logic_flag=business_flag
            )
        
        # Store is in the report, check node specifics
        offline_nodes_for_store = self.offline_nodes[ticket.store_number]
        
        if ticket.node_number is None:
            # Can't determine specific node - needs review
            confidence = self.determine_confidence(ticket, True, False, business_flag)
            return self.create_analysis_result(
                ticket=ticket,
                status="needs_review",
                reason=self._format_offline_nodes_message(offline_nodes_for_store, "but couldn't identify specific node from description"),
                store_in_report=True,
                node_in_report=False,
                confidence=confidence,
                business_flag=business_flag,
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        # Check if the specific node is offline
        node_in_report = ticket.node_number in offline_nodes_for_store
        confidence = self.determine_confidence(ticket, store_in_report, node_in_report, business_flag)
        
        if node_in_report:
            # The specific node is indeed offline - needs review
            return self.create_analysis_result(
                ticket=ticket,
                status="needs_review",
                reason=f"Node {ticket.node_number} is confirmed offline in the report",
                store_in_report=True,
                node_in_report=True,
                confidence=confidence,
                business_flag=business_flag,
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        else:
            # The specific node is NOT offline - ticket can be closed
            return AnalysisResult(
                ticket=ticket,
                status="can_close",
                reason=f"Node {ticket.node_number} is not in offline report. Offline nodes for store: {sorted(offline_nodes_for_store)}",
                store_in_report=True,
                node_in_report=False,
                confidence=confidence,
                business_logic_flag=business_flag
            )
    
    def analyze_closed_ticket(self, ticket: Ticket) -> AnalysisResult:
        """Analyze a closed ticket to determine if it should be reopened"""
        
        # Check if we could extract store number
        if ticket.store_number is None:
            return self.create_analysis_result(
                ticket=ticket,
                status="error",
                reason="Could not extract store number from site field",
                store_in_report=False,
                node_in_report=False,
                confidence="low",
                business_flag=""
            )
        
        # Analyze temporal correlation first
        temporal_analysis, days_offline, has_timeline_issue = self.analyze_temporal_correlation(ticket)
        
        # Check if store is in the offline report
        store_in_report = ticket.store_number in self.offline_nodes
        
        if not store_in_report:
            # Store not in report means no nodes are offline - ticket should stay closed
            return self.create_analysis_result(
                ticket=ticket,
                status="closed_ok",
                reason="Store not in offline report - all nodes are online, ticket correctly closed",
                store_in_report=False,
                node_in_report=False,
                confidence="high",
                business_flag="",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        # Store has offline nodes - check if this ticket should be reopened
        offline_nodes_for_store = self.offline_nodes[ticket.store_number]
        
        # Check for CRITICAL conditions first
        is_saf_store = ticket.store_number in self.saf_stores
        is_both_nodes_offline = ticket.store_number in self.both_nodes_offline_stores
        
        if is_saf_store:
            return self.create_analysis_result(
                ticket=ticket,
                status="suggest_reopen",
                reason=f"CRITICAL: Store has SAF (Store and Forward) failure - ticket should be REOPENED IMMEDIATELY. " + self._format_offline_nodes_message(offline_nodes_for_store, ""),
                store_in_report=True,
                node_in_report=True,
                confidence="high",
                business_flag="critical_saf",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        if is_both_nodes_offline:
            return self.create_analysis_result(
                ticket=ticket,
                status="suggest_reopen",
                reason=f"CRITICAL: Store has BOTH nodes offline - ticket should be REOPENED IMMEDIATELY. " + self._format_offline_nodes_message(offline_nodes_for_store, ""),
                store_in_report=True,
                node_in_report=True,
                confidence="high",
                business_flag="critical_both_nodes_offline",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        # Check specific node if we can identify it
        if ticket.node_number is None:
            # Can't determine specific node, but store has offline nodes - suggest review
            return self.create_analysis_result(
                ticket=ticket,
                status="suggest_reopen",
                reason=self._format_offline_nodes_message(offline_nodes_for_store, "but couldn't identify specific node from ticket description - SUGGEST REVIEW FOR REOPEN"),
                store_in_report=True,
                node_in_report=False,
                confidence="medium",
                business_flag="",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        
        # Check if the specific node mentioned in the ticket is offline
        node_in_report = ticket.node_number in offline_nodes_for_store
        
        if node_in_report:
            # The specific node is offline - ticket should be reopened
            return self.create_analysis_result(
                ticket=ticket,
                status="suggest_reopen",
                reason=f"Node {ticket.node_number} is confirmed offline - ticket should be REOPENED",
                store_in_report=True,
                node_in_report=True,
                confidence="high",
                business_flag="",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
        else:
            # The specific node is online, but other nodes are offline - ticket stays closed
            return self.create_analysis_result(
                ticket=ticket,
                status="closed_ok",
                reason=f"Node {ticket.node_number} is online (correctly closed). Other " + self._format_offline_nodes_message(offline_nodes_for_store).lower(),
                store_in_report=True,
                node_in_report=False,
                confidence="high",
                business_flag="",
                temporal_analysis=temporal_analysis,
                days_offline=days_offline
            )
    
    def analyze_all_tickets(self):
        """Analyze all tickets (both open and closed)"""
        print("Analyzing tickets...")
        
        open_tickets = [t for t in self.tickets if not t.is_closed]
        closed_tickets = [t for t in self.tickets if t.is_closed]
        
        print(f"  Processing {len(open_tickets)} open tickets...")
        for ticket in open_tickets:
            result = self.analyze_ticket(ticket)
            self.results.append(result)
        
        if closed_tickets:
            print(f"  Processing {len(closed_tickets)} closed tickets...")
            for ticket in closed_tickets:
                result = self.analyze_closed_ticket(ticket)
                self.results.append(result)
        
        # Print summary
        can_close = len([r for r in self.results if r.status == "can_close"])
        needs_review = len([r for r in self.results if r.status == "needs_review"])
        suggest_reopen = len([r for r in self.results if r.status == "suggest_reopen"])
        closed_ok = len([r for r in self.results if r.status == "closed_ok"])
        closed_too_old = len([r for r in self.results if r.status == "closed_too_old"])
        errors = len([r for r in self.results if r.status == "error"])
        
        # Check for missing tickets
        missing_tickets = self.get_missing_tickets()
        
        print(f"Analysis complete:")
        print(f"  Can close: {can_close}")
        print(f"  Need review: {needs_review}")
        print(f"  Suggest reopen: {suggest_reopen}")
        print(f"  Closed OK: {closed_ok}")
        print(f"  Closed too old (>7 days): {closed_too_old}")
        print(f"  Errors: {errors}")
        print(f"  Missing tickets: {len(missing_tickets)} (stores with offline nodes but no tickets)")
    
    def export_results(self):
        """Export results to CSV files, Excel file, and summary report"""
        
        # Define common headers for all CSV files (matching Excel)
        csv_headers = [
            'Ticket_Number', 'Site', 'Description', 'Priority', 
            'Created', 'Updated', 'Resolved', 'Assignment_Group', 'Ticket_Status', 'Store_Number', 'Node_Number', 
            'Days_Offline', 'Reopenable', 'Confidence', 'Business_Flag', 'Temporal_Analysis', 'Reason'
        ]
        
        # Export tickets that can be closed
        can_close_tickets = [r for r in self.results if r.status == "can_close"]
        if can_close_tickets:
            with open('results_can_close.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(csv_headers)
                
                for result in can_close_tickets:
                    t = result.ticket
                    writer.writerow([
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ])
            print(f"Exported {len(can_close_tickets)} closable tickets to results_can_close.csv")
        
        # Export tickets that need review
        needs_review_tickets = [r for r in self.results if r.status == "needs_review"]
        if needs_review_tickets:
            with open('results_need_review.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(csv_headers)
                
                for result in needs_review_tickets:
                    t = result.ticket
                    writer.writerow([
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ])
            print(f"Exported {len(needs_review_tickets)} tickets needing review to results_need_review.csv")
        
        # Export tickets that suggest reopening
        suggest_reopen_tickets = [r for r in self.results if r.status == "suggest_reopen"]
        if suggest_reopen_tickets:
            with open('results_suggest_reopen.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(csv_headers)
                
                for result in suggest_reopen_tickets:
                    t = result.ticket
                    writer.writerow([
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ])
            print(f"Exported {len(suggest_reopen_tickets)} closed tickets suggesting reopen to results_suggest_reopen.csv")
        
        # Export tickets that are correctly closed
        closed_ok_tickets = [r for r in self.results if r.status == "closed_ok"]
        if closed_ok_tickets:
            with open('results_closed_ok.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(csv_headers)
                
                for result in closed_ok_tickets:
                    t = result.ticket
                    writer.writerow([
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ])
            print(f"Exported {len(closed_ok_tickets)} correctly closed tickets to results_closed_ok.csv")
        
        # Export errors if any
        error_tickets = [r for r in self.results if r.status == "error"]
        if error_tickets:
            with open('results_errors.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(csv_headers)
                
                for result in error_tickets:
                    t = result.ticket
                    writer.writerow([
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ])
            print(f"Exported {len(error_tickets)} error tickets to results_errors.csv")
        
        # Export to Excel if available
        if EXCEL_AVAILABLE:
            missing_tickets = self.get_missing_tickets()
            closed_too_old_tickets = [r for r in self.results if r.status == "closed_too_old"]
            self.export_to_excel(can_close_tickets, needs_review_tickets, suggest_reopen_tickets, closed_ok_tickets, closed_too_old_tickets, error_tickets, missing_tickets)
        else:
            print("Excel export unavailable - openpyxl not installed. Run: pip install openpyxl")
        
        # Create summary report
        self.create_summary_report()
    
    def export_to_excel(self, can_close_tickets, needs_review_tickets, suggest_reopen_tickets, closed_ok_tickets, closed_too_old_tickets, error_tickets, missing_tickets):
        """Export results to Excel workbook with multiple sheets"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            can_close_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            headers = [
                'Ticket_Number', 'Site', 'Description', 'Priority', 
                'Created', 'Updated', 'Resolved', 'Assignment_Group', 'Ticket_Status', 'Store_Number', 'Node_Number', 
                'Days_Offline', 'Reopenable', 'Confidence', 'Business_Flag', 'Temporal_Analysis', 'Reason'
            ]
            
            # Sheet 1: Can Close
            if can_close_tickets:
                ws_close = wb.create_sheet("Can Close")
                ws_close.append(headers)
                
                # Style header row
                for col_num, header in enumerate(headers, 1):
                    cell = ws_close.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data rows
                for result in can_close_tickets:
                    t = result.ticket
                    row = [
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ]
                    ws_close.append(row)
                    
                    # Highlight high confidence rows
                    if result.confidence == "high":
                        row_num = ws_close.max_row
                        for col_num in range(1, len(headers) + 1):
                            ws_close.cell(row=row_num, column=col_num).fill = can_close_fill
                
                # Auto-adjust column widths
                for column in ws_close.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_close.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Need Review
            if needs_review_tickets:
                ws_review = wb.create_sheet("Need Review")
                ws_review.append(headers)
                
                # Style header row
                for col_num, header in enumerate(headers, 1):
                    cell = ws_review.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data rows
                for result in needs_review_tickets:
                    t = result.ticket
                    row = [
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ]
                    ws_review.append(row)
                    
                    # Highlight critical rows
                    row_num = ws_review.max_row
                    if result.business_logic_flag in ["critical_saf", "critical_both_nodes_offline"]:
                        for col_num in range(1, len(headers) + 1):
                            ws_review.cell(row=row_num, column=col_num).fill = critical_fill
                            ws_review.cell(row=row_num, column=col_num).font = Font(color="FFFFFF", bold=True)
                    elif result.confidence == "high":
                        for col_num in range(1, len(headers) + 1):
                            ws_review.cell(row=row_num, column=col_num).fill = high_fill
                
                # Auto-adjust column widths
                for column in ws_review.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_review.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 3: Suggest Reopen
            if suggest_reopen_tickets:
                ws_reopen = wb.create_sheet("Suggest Reopen")
                ws_reopen.append(headers)
                
                # Style header row
                for col_num, header in enumerate(headers, 1):
                    cell = ws_reopen.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data rows
                for result in suggest_reopen_tickets:
                    t = result.ticket
                    row = [
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ]
                    ws_reopen.append(row)
                    
                    # Highlight critical rows
                    row_num = ws_reopen.max_row
                    if result.business_logic_flag in ["critical_saf", "critical_both_nodes_offline"]:
                        for col_num in range(1, len(headers) + 1):
                            ws_reopen.cell(row=row_num, column=col_num).fill = critical_fill
                            ws_reopen.cell(row=row_num, column=col_num).font = Font(color="FFFFFF", bold=True)
                    elif result.confidence == "high":
                        for col_num in range(1, len(headers) + 1):
                            ws_reopen.cell(row=row_num, column=col_num).fill = high_fill
                
                # Auto-adjust column widths
                for column in ws_reopen.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_reopen.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 4: Closed OK
            if closed_ok_tickets:
                ws_closed_ok = wb.create_sheet("Closed OK")
                ws_closed_ok.append(headers)
                
                # Style header row
                for col_num, header in enumerate(headers, 1):
                    cell = ws_closed_ok.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data rows
                for result in closed_ok_tickets:
                    t = result.ticket
                    row = [
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ]
                    ws_closed_ok.append(row)
                    
                    # Highlight high confidence rows with light green
                    if result.confidence == "high":
                        row_num = ws_closed_ok.max_row
                        for col_num in range(1, len(headers) + 1):
                            ws_closed_ok.cell(row=row_num, column=col_num).fill = can_close_fill
                
                # Auto-adjust column widths
                for column in ws_closed_ok.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_closed_ok.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 5: Errors (if any)
            if error_tickets:
                ws_errors = wb.create_sheet("Errors")
                ws_errors.append(headers)
                
                # Style header row
                for col_num, header in enumerate(headers, 1):
                    cell = ws_errors.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data rows
                for result in error_tickets:
                    t = result.ticket
                    row = [
                        t.number, t.site, t.description, t.priority,
                        t.created, t.updated, t.resolved or '', t.assignment_group or '', 'OPEN' if not t.is_closed else 'CLOSED', t.store_number, t.node_number,
                        result.days_offline, 'Yes' if result.reopenable else 'No', result.confidence, result.business_logic_flag, result.temporal_analysis, result.reason
                    ]
                    ws_errors.append(row)
                
                # Auto-adjust column widths
                for column in ws_errors.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_errors.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 4: Missing Tickets
            if missing_tickets:
                ws_missing = wb.create_sheet("Missing Tickets")
                missing_headers = [
                    'Store_Number', 'Site', 'Node_Number', 'Priority', 'Urgency',
                    'Suggested_Description', 'All_Offline_Nodes', 'SAF_Store', 'Both_Nodes_Offline', 'Reason'
                ]
                ws_missing.append(missing_headers)
                
                # Style header row
                for col_num, header in enumerate(missing_headers, 1):
                    cell = ws_missing.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data rows
                for missing in missing_tickets:
                    row = [
                        missing['store_number'],
                        missing['site'],
                        missing['node_number'],
                        missing['priority'],
                        missing['urgency'],
                        missing['suggested_description'],
                        ', '.join(map(str, missing['offline_nodes'])),
                        'YES' if missing['is_saf'] else 'NO',
                        'YES' if missing['is_both_offline'] else 'NO',
                        missing['reason']
                    ]
                    ws_missing.append(row)
                    
                    # Highlight critical rows
                    row_num = ws_missing.max_row
                    if missing['is_saf'] or missing['is_both_offline']:
                        for col_num in range(1, len(missing_headers) + 1):
                            ws_missing.cell(row=row_num, column=col_num).fill = critical_fill
                            ws_missing.cell(row=row_num, column=col_num).font = Font(color="FFFFFF", bold=True)
                    elif missing['priority'] == "High":
                        for col_num in range(1, len(missing_headers) + 1):
                            ws_missing.cell(row=row_num, column=col_num).fill = high_fill
                
                # Auto-adjust column widths
                for column in ws_missing.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_missing.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 5: Summary
            ws_summary = wb.create_sheet("Summary")
            
            # Summary statistics
            total_tickets = len(self.results)
            can_close_count = len(can_close_tickets)
            needs_review_count = len(needs_review_tickets)
            errors_count = len(error_tickets)
            
            missing_count = len(missing_tickets)
            summary_data = [
                ["Node Cross-Reference Analysis Summary", ""],
                ["", ""],
                ["Analysis Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["", ""],
                ["OVERALL STATISTICS", ""],
                ["Total tickets analyzed", total_tickets],
                ["Can close", f"{can_close_count} ({can_close_count/total_tickets*100:.1f}%)"],
                ["Need review", f"{needs_review_count} ({needs_review_count/total_tickets*100:.1f}%)"],
                ["Errors", f"{errors_count} ({errors_count/total_tickets*100:.1f}%)"],
                ["Missing tickets needed", missing_count],
                ["", ""],
                ["CRITICAL CONDITIONS", ""],
                ["Stores with SAF markers", len(self.saf_stores)],
                ["Stores with both nodes offline", len(self.both_nodes_offline_stores)],
                ["SAF Store Numbers", ", ".join(map(str, sorted(self.saf_stores))) if self.saf_stores else "None"],
                ["Both Nodes Offline", ", ".join(map(str, sorted(self.both_nodes_offline_stores))) if self.both_nodes_offline_stores else "None"],
                ["", ""],
                ["PROACTIVE MONITORING", ""],
                ["Stores needing new tickets", len(set(mt['store_number'] for mt in missing_tickets))],
                ["Critical missing tickets", len([mt for mt in missing_tickets if 'CRITICAL' in mt['priority']])],
                ["High priority missing tickets", len([mt for mt in missing_tickets if mt['priority'] == 'High'])]
            ]
            
            for row_data in summary_data:
                ws_summary.append(row_data)
            
            # Style summary sheet
            ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
            ws_summary.cell(row=5, column=1).font = Font(bold=True)
            ws_summary.cell(row=11, column=1).font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in ws_summary.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            filename = f"node_cross_reference_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            print(f"Excel report exported to: {filename}")
            print(f"  - {'Can Close' if can_close_tickets else 'No Can Close'} sheet: {len(can_close_tickets)} tickets")
            print(f"  - {'Need Review' if needs_review_tickets else 'No Need Review'} sheet: {len(needs_review_tickets)} tickets")
            print(f"  - {'Suggest Reopen' if suggest_reopen_tickets else 'No Suggest Reopen'} sheet: {len(suggest_reopen_tickets)} tickets")
            print(f"  - {'Closed OK' if closed_ok_tickets else 'No Closed OK'} sheet: {len(closed_ok_tickets)} tickets") 
            print(f"  - {'Errors' if error_tickets else 'No Errors'} sheet: {len(error_tickets)} tickets")
            print(f"  - {'Missing Tickets' if missing_tickets else 'No Missing Tickets'} sheet: {len(missing_tickets)} tickets needed")
            print(f"  - Summary sheet with critical conditions overview")
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
    
    def create_summary_report(self):
        """Create a detailed summary report"""
        with open('summary_report.txt', 'w', encoding='utf-8') as f:
            f.write("NODE CROSS-REFERENCE ANALYSIS SUMMARY\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Analysis performed on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # Overall statistics
            total_tickets = len(self.results)
            can_close = len([r for r in self.results if r.status == "can_close"])
            needs_review = len([r for r in self.results if r.status == "needs_review"])
            suggest_reopen = len([r for r in self.results if r.status == "suggest_reopen"])
            closed_ok = len([r for r in self.results if r.status == "closed_ok"])
            errors = len([r for r in self.results if r.status == "error"])
            
            # Confidence breakdown
            high_conf = len([r for r in self.results if r.confidence == "high"])
            med_conf = len([r for r in self.results if r.confidence == "medium"])
            low_conf = len([r for r in self.results if r.confidence == "low"])
            
            # Business logic flags
            business_flagged = len([r for r in self.results if r.business_logic_flag])
            
            f.write("OVERALL STATISTICS:\n")
            f.write(f"Total tickets analyzed: {total_tickets}\n")
            f.write(f"Can close: {can_close} ({can_close/total_tickets*100:.1f}%)\n")
            f.write(f"Need review: {needs_review} ({needs_review/total_tickets*100:.1f}%)\n")
            f.write(f"Suggest reopen: {suggest_reopen} ({suggest_reopen/total_tickets*100:.1f}%)\n")
            f.write(f"Closed OK: {closed_ok} ({closed_ok/total_tickets*100:.1f}%)\n")
            f.write(f"Errors: {errors} ({errors/total_tickets*100:.1f}%)\n\n")
            
            f.write("CONFIDENCE BREAKDOWN:\n")
            f.write(f"High confidence: {high_conf} ({high_conf/total_tickets*100:.1f}%)\n")
            f.write(f"Medium confidence: {med_conf} ({med_conf/total_tickets*100:.1f}%)\n")
            f.write(f"Low confidence: {low_conf} ({low_conf/total_tickets*100:.1f}%)\n\n")
            
            f.write("BUSINESS LOGIC FLAGS:\n")
            f.write(f"Tickets with business logic flags: {business_flagged} ({business_flagged/total_tickets*100:.1f}%)\n")
            if business_flagged > 0:
                flag_counts = {}
                for result in self.results:
                    if result.business_logic_flag:
                        flag_counts[result.business_logic_flag] = flag_counts.get(result.business_logic_flag, 0) + 1
                for flag, count in flag_counts.items():
                    f.write(f"  {flag}: {count} tickets\n")
            f.write("\n")
            
            # Breakdown by reason
            f.write("BREAKDOWN BY REASON:\n")
            reason_counts = {}
            for result in self.results:
                key = f"{result.status}: {result.reason.split('.')[0]}"  # First sentence only
                reason_counts[key] = reason_counts.get(key, 0) + 1
            
            for reason, count in sorted(reason_counts.items(), key=lambda x: x[1], reverse=True):
                f.write(f"  {reason}: {count} tickets\n")
            
            f.write("\n")
            
            # Store analysis
            f.write("STORE ANALYSIS:\n")
            stores_with_tickets = set()
            stores_in_report = set()
            
            for result in self.results:
                if result.ticket.store_number:
                    stores_with_tickets.add(result.ticket.store_number)
                    if result.store_in_report:
                        stores_in_report.add(result.ticket.store_number)
            
            f.write(f"Unique stores with tickets: {len(stores_with_tickets)}\n")
            f.write(f"Stores with tickets that have offline nodes: {len(stores_in_report)}\n")
            f.write(f"Stores with tickets that have NO offline nodes: {len(stores_with_tickets - stores_in_report)}\n\n")
            
            # Output files generated
            f.write("OUTPUT FILES GENERATED:\n")
            if can_close > 0:
                f.write("  - results_can_close.csv: Tickets that can be definitively closed\n")
            if needs_review > 0:
                f.write("  - results_need_review.csv: Tickets requiring manual review\n")
            if suggest_reopen > 0:
                f.write("  - results_suggest_reopen.csv: Closed tickets that should be reopened\n")
            if closed_ok > 0:
                f.write("  - results_closed_ok.csv: Closed tickets that should stay closed\n")
            if errors > 0:
                f.write("  - results_errors.csv: Tickets with parsing errors\n")
            f.write("  - summary_report.txt: This summary report\n")
            if EXCEL_AVAILABLE:
                f.write("  - Excel workbook with all categories and summary\n")
        
        print("Summary report created: summary_report.txt")


def main():
    """Main function to run the cross-reference analysis"""
    
    # Check if input files exist
    csv_file = "newnode.csv"  # Using new larger dataset
    report_file = "data/nodes_report.txt"
    
    if not os.path.exists(csv_file):
        print(f"Error: {csv_file} not found in current directory")
        return
    
    if not os.path.exists(report_file):
        print(f"Error: {report_file} not found in current directory")
        return
    
    print("Node Cross-Reference Analysis Tool")
    print("=" * 40)
    print()
    
    # Initialize and run analysis
    cross_ref = NodeCrossReference()
    
    try:
        # Load data
        cross_ref.load_tickets(csv_file)
        cross_ref.load_offline_nodes(report_file)
        
        # Analyze
        cross_ref.analyze_all_tickets()
        
        # Export results
        cross_ref.export_results()

        # Clean up temporary files
        cross_ref.cleanup()

        print("\nAnalysis complete! Check the output files for results.")
        
    except Exception as e:
        print(f"Error during analysis: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()